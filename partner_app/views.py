from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.core.files.base import ContentFile
from django.forms.models import model_to_dict
from django.core.mail import send_mail, EmailMessage
from django.utils import timezone
from django.db.models import Sum, Count, Avg, F, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncDate
from django.contrib import messages
from django.shortcuts import redirect
import logging
from django.contrib.auth import update_session_auth_hash
from datetime import datetime, timedelta
import os
import json
import tempfile
from moviepy import VideoFileClip
from PIL import Image
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.core.exceptions import ObjectDoesNotExist
from core.models import (
    Event,
    Ticket,
    Order,
    PayoutRequest,
    PayoutDetails,
    Tag,
    MainTag,
    EventPackage,
    UserPackageSubscription,
)
from .forms import EventForm, DocumentUploadForm, ReportScheduleForm, PayoutDetailsForm
from .models import SalesReport, ReportSchedule
from .utils import generate_sales_report
from core.forms import PartnerProfileForm, PasswordChangeForm
from django.utils import timezone

logger = logging.getLogger(__name__)

def get_rejection_messages(request):
    """Возвращает сообщения об отклонении мероприятий для текущего пользователя."""
    from core.models import Event

    rejected_events = Event.objects.filter(organizer=request.user, status="rejected")

    rejection_messages = []
    for event in rejected_events:
        if event.rejection_reason:
            rejection_messages.append(
                f"Мероприятие {event.title} отклонено. Причина: {event.rejection_reason}"
            )

    return rejection_messages

@login_required
def partner_dashboard(request):
    if request.user.user_type != "partner":
        return redirect("visitor:dashboard")

    # Получаем отклоненные мероприятия партнёра
    rejected_events = Event.objects.filter(organizer=request.user, status="rejected")

    rejection_messages = []
    for event in rejected_events:
        if event.rejection_reason:
            rejection_messages.append(
                f"Мероприятие '{event.title}' отклонено. Причина: {event.rejection_reason}"
            )

    # Получаем активные мероприятия партнёра
    active_events = Event.objects.filter(
        organizer=request.user, status="active"
    ).count()

    # Получаем продажи за текущий месяц
    from datetime import datetime

    current_month = datetime.now().month
    current_year = datetime.now().year

    monthly_sales = (
        Order.objects.filter(
            ticket__event__organizer=request.user,
            created_at__year=current_year,
            created_at__month=current_month,
        ).aggregate(total=Sum("total_price"))["total"]
        or 0
    )

    # Получаем ожидающие выплаты
    pending_payouts = PayoutRequest.objects.filter(
        organizer=request.user, status="pending"
    ).count()

    # Получаем активную подписку пользователя
    user_subscription = (
        UserPackageSubscription.objects.filter(user=request.user, is_active=True)
        .select_related("package")
        .first()
    )

    context = {
        "user": request.user,
        "active_events_count": active_events,
        "monthly_sales_sum": monthly_sales,
        "pending_payouts_count": pending_payouts,
        "rejection_messages": rejection_messages,
        "packages": EventPackage.objects.all(),
        "user_subscription": user_subscription,
    }
    return render(request, "partner/dashboard.html", context)

@login_required
def create_event(request):
    """
    View для создания нового мероприятия.
    Видео обрабатывается автоматически через сигналы и Celery.
    """

    # Проверяем наличие активного пакета у пользователя
    active_subscription = (
        UserPackageSubscription.objects.filter(
            user=request.user,
            is_active=True,
        )
        .select_related("package")
        .first()
    )

    if request.method == "POST":
        # Если это редактирование существующего мероприятия
        if 'event_id' in request.POST:
            event = Event.objects.filter(id=request.POST['event_id'], organizer=request.user).first()
            if event and event.package:
                # Если у события уже есть пакет, используем его
                package = event.package
            else:
                # Если нет активного пакета - перенаправляем на покупку
                if not active_subscription:
                    messages.error(
                        request,
                        "Для создания мероприятий необходимо приобрести пакет. Пожалуйста, выберите и оплатите подходящий пакет."
                    )
                    return redirect("payment:package_selection")
                package = active_subscription.package
        else:
            # Если это создание нового мероприятия
            if not active_subscription:
                messages.error(
                    request,
                    "Для создания мероприятий необходимо приобрести пакет. Пожалуйста, выберите и оплатите подходящий пакет."
                )
                return redirect("payment:package_selection")
            package = active_subscription.package

        form = EventForm(request.POST, request.FILES, user=request.user, current_package=package)
        if form.is_valid():
            # Проверяем типы загружаемых файлов
            video_file = request.FILES.get("video_url")
            if video_file:
                # Проверяем расширение видео
                valid_video_extensions = ['.mp4', '.mov', '.avi']
                if not any(video_file.name.lower().endswith(ext) for ext in valid_video_extensions):
                    messages.error(
                        request,
                        "Неверный формат видео. Разрешены только файлы MP4, MOV, AVI"
                    )
                    return render(
                        request,
                        "partner/event_form.html",
                        {
                            "form": form,
                            "is_edit": False,
                            "ticket_data": [],
                            "rejection_messages": get_rejection_messages(request),
                            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                            "has_free_tickets": False,
                            "packages": EventPackage.objects.all(),
                        },
                    )

# Проверяем длительность видео
                try:
                    temp_file_path = None
                    temp_file_path = tempfile.mktemp(suffix=os.path.splitext(video_file.name)[1])
                    
                    # Сохраняем видео во временный файл
                    with open(temp_file_path, 'wb+') as temp_file:
                        for chunk in video_file.chunks():
                            temp_file.write(chunk)

                    # Проверяем длительность
                    with VideoFileClip(temp_file_path) as video:
                        duration = video.duration
                        max_duration = 310  # 5 минут в секундах
                        if duration > max_duration:
                            if os.path.exists(temp_file_path):
                                os.unlink(temp_file_path)
                            messages.error(
                                request,
                                "Длительность видео превышает 5 минут. Пожалуйста, загрузите видео не длиннее 5 минут."
                            )
                            return render(
                                request,
                                "partner/event_form.html",
                                {
                                    "form": form,
                                    "is_edit": False,
                                    "ticket_data": [],
                                    "rejection_messages": get_rejection_messages(request),
                                    "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                                    "has_free_tickets": False,
                                    "packages": EventPackage.objects.all(),
                                },
                            )

                    # Удаляем временный файл
                    if os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
                    
                    # Восстанавливаем указатель файла после проверки
                    video_file.file.seek(0)
                except Exception as e:
                    logger.error(f"Ошибка при проверке длительности видео: {str(e)}")
                    # Очищаем временный файл при ошибке
                    if temp_file_path and os.path.exists(temp_file_path):
                        try:
                            os.unlink(temp_file_path)
                        except:
                            pass
                    messages.error(
                        request,
                        "Произошла ошибка при проверке длительности видео. Пожалуйста, попробуйте еще раз."
                    )
                    return render(
                        request,
                        "partner/event_form.html",
                        {
                            "form": form,
                            "is_edit": False,
                            "ticket_data": [],
                            "rejection_messages": get_rejection_messages(request),
                            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                            "has_free_tickets": False,
                            "packages": EventPackage.objects.all(),
                        },
                    )

            pdf_file = request.FILES.get("program_file")
            if pdf_file:
                # Проверяем расширение PDF
                if not pdf_file.name.lower().endswith('.pdf'):
                    messages.error(
                        request,
                        "Неверный формат файла. Разрешены только PDF файлы"
                    )
                    return render(
                        request,
                        "partner/event_form.html",
                        {
                            "form": form,
                            "is_edit": False,
                            "ticket_data": [],
                            "rejection_messages": get_rejection_messages(request),
                            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                            "has_free_tickets": False,
                            "packages": EventPackage.objects.all(),
                        },
                    )

            # Проверяем типы загружаемых изображений
            main_image = request.FILES.get("image")
            additional_images = request.FILES.getlist("images")

            # Валидация основного изображения
            if main_image:
                valid_image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
                if not any(main_image.name.lower().endswith(ext) for ext in valid_image_extensions):
                    messages.error(
                        request,
                        "Неверный формат основного изображения. Разрешены только JPG, PNG, GIF, WEBP"
                    )
                    return render(
                        request,
                        "partner/event_form.html",
                        {
                            "form": form,
                            "is_edit": False,
                            "ticket_data": [
                                {"name": name, "price": price, "quantity": quantity}
                                for name, price, quantity in zip(
                                    request.POST.getlist("ticket_name[]"),
                                    request.POST.getlist("ticket_price[]"),
                                    request.POST.getlist("ticket_quantity[]"),
                                )
                                if name and price and quantity
                            ],
                            "rejection_messages": get_rejection_messages(request),
                            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                            "has_free_tickets": False,
                            "packages": EventPackage.objects.all(),
                        },
                    )

            # Валидация дополнительных изображений
            if additional_images:
                valid_image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
                for image in additional_images:
                    if not any(image.name.lower().endswith(ext) for ext in valid_image_extensions):
                        messages.error(
                            request,
                            "Неверный формат дополнительного изображения. Разрешены только JPG, PNG, GIF, WEBP"
                        )
                        return render(
                            request,
                            "partner/event_form.html",
                            {
                                "form": form,
                                "is_edit": False,
                                "ticket_data": [
                                    {"name": name, "price": price, "quantity": quantity}
                                    for name, price, quantity in zip(
                                        request.POST.getlist("ticket_name[]"),
                                        request.POST.getlist("ticket_price[]"),
                                        request.POST.getlist("ticket_quantity[]"),
                                    )
                                    if name and price and quantity
                                ],
                                "rejection_messages": get_rejection_messages(request),
                                "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                                "has_free_tickets": False,
                                "packages": EventPackage.objects.all(),
                            },
                        )

            # Считаем общее количество загружаемых фото
            total_images = 0
            if main_image:
                total_images += 1
            if additional_images:
                total_images += len(additional_images)

            # Проверяем не превышаем ли лимит пакета
            if total_images > package.max_photos:
                additional_photos_allowed = package.max_photos - 1
                messages.error(
                    request,
                    f"Ваш пакет позволяет загрузить не более {package.max_photos} фотографий "
                    f"(1 основное + {additional_photos_allowed} дополнительных). "
                    f"Вы пытаетесь загрузить {total_images} фото."
                )
                return render(
                    request,
                    "partner/event_form.html",
                    {
                        "form": form,
                        "is_edit": False,
                        "ticket_data": [],
                        "rejection_messages": get_rejection_messages(request),
                        "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                        "has_free_tickets": False,
                        "packages": EventPackage.objects.all(),
                    },
                )

            event = form.save(commit=False)
            event.organizer = request.user
            event.status = "on_moderation"

            # Очищаем медиафайлы, если были отмечены соответствующие флаги
            for field_name in ["video_url", "program_file"]:
                clear_field_name = f"{field_name}-clear"
                if clear_field_name in request.POST:
                    current_file = getattr(event, field_name)
                    if current_file:
                        current_file.delete(save=False)  # Не сохраняем модель здесь
                    setattr(event, field_name, None)

            # Основное фото (отдельный input image), если он есть в запросе
            main_image = request.FILES.get("image")
            if main_image:
                event.image = main_image

            event.save()

            # Дополнительные фото (много, input images)
            images = request.FILES.getlist("images")
            if images:
                from core.models import EventImage
                for image in images:
                    EventImage.objects.create(event=event, image=image)

        else:
            # если форма не валидна — не создаём/сохраняем event здесь
            event = None

        if event is None:
            # просто отдадим форму как есть (предыдущая логика уже делает ticket_data для невалидного POST)
            form = EventForm(request.POST, request.FILES, user=request.user, current_package=active_subscription.package)
            ticket_data = [
                {"name": name, "price": price, "quantity": quantity}
                for name, price, quantity in zip(
                    request.POST.getlist("ticket_name[]"),
                    request.POST.getlist("ticket_price[]"),
                    request.POST.getlist("ticket_quantity[]"),
                )
                if name and price and quantity
            ]
            return render(
                request,
                "partner/event_form.html",
                {
                    "form": form,
                    "is_edit": False,
                    "ticket_data": ticket_data,
                    "rejection_messages": get_rejection_messages(request),
                    "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                    "has_free_tickets": False,
                    "packages": EventPackage.objects.all(),
                },
            )

        # Обрабатываем теги из массива ID
        tags_ids = request.POST.getlist("tags")
        if tags_ids:
            # Ограничиваем количество тегов до 5
            selected_tags = tags_ids[:5]
            event.tags.set(selected_tags)

        # Обрабатываем данные о билетах из таблицы
        ticket_names = request.POST.getlist("ticket_name[]")
        ticket_prices = request.POST.getlist("ticket_price[]")
        ticket_quantities = request.POST.getlist("ticket_quantity[]")

        # Проверяем, есть ли одновременно бесплатные и платные билеты
        has_free_tickets = False
        has_paid_tickets = False

        for price in ticket_prices:
            if price:
                try:
                    price_value = (
                        float(price.replace(",", ".")) if "," in price else float(price)
                    )
                    if price_value == 0:
                        has_free_tickets = True
                    else:
                        has_paid_tickets = True
                except (ValueError, TypeError):
                    continue

        # Если есть и бесплатные, и платные билеты одновременно
        if has_free_tickets and has_paid_tickets:
            messages.error(
                request,
                "Невозможно создать мероприятие с бесплатными и платными билетами одновременно.",
            )
            return render(
                request,
                "partner/event_form.html",
                {
                    "form": form,
                    "is_edit": False,
                    "ticket_data": [
                        {"name": name, "price": price, "quantity": quantity}
                        for name, price, quantity in zip(
                            ticket_names, ticket_prices, ticket_quantities
                        )
                        if name and price and quantity
                    ],
                    "rejection_messages": get_rejection_messages(request),
                    "all_tags": Tag.objects.all(),
                },
            )

        for name, price, quantity in zip(
            ticket_names, ticket_prices, ticket_quantities
        ):
            if name and price and quantity:
                try:
                    event.tickets.create(
                        name=name,
                        price=(
                            float(price.replace(",", "."))
                            if "," in price
                            else float(price)
                        ),
                        available_quantity=int(quantity),
                    )
                except (ValueError, TypeError):
                    continue

        # РЕДИРЕКТИМ пользователя.
        # Обработка видео начнется автоматически через сигнал post_save.
        messages.success(
            request,
            "Мероприятие успешно создано! Видео будет обработано в фоновом режиме.",
        )
        return redirect("partner:partner_event_list")
    else:
        # При загрузке страницы (GET) выводим пакет, который сейчас активен по подписке пользователя
        if not active_subscription:
            messages.warning(
                request,
                "У вас нет активного пакета. Пожалуйста, выберите и купите пакет для создания мероприятий."
            )
            return redirect("payment:package_selection")

        form = EventForm(user=request.user, current_package=active_subscription.package)

    ticket_data = []
    if request.method == "POST" and not form.is_valid():
        ticket_data = [
            {"name": name, "price": price, "quantity": quantity}
            for name, price, quantity in zip(
                request.POST.getlist("ticket_name[]"),
                request.POST.getlist("ticket_price[]"),
                request.POST.getlist("ticket_quantity[]"),
            )
            if name and price and quantity
        ]

    return render(
        request,
        "partner/event_form.html",
        {
            "form": form,
            "is_edit": False,
            "ticket_data": ticket_data,
            "rejection_messages": get_rejection_messages(request),
            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
            "has_free_tickets": False,  # По умолчанию False, будет обновляться через JavaScript
            "packages": EventPackage.objects.all(),
        },
    )

def notify_organizer(event):
    subject = f"Ваше мероприятие '{event.title}' одобрено!"
    message = f"Привет, {event.organizer.first_name}!\n\nВаше мероприятие '{event.title}' успешно добавлено на сайт."
    send_mail(subject, message, "dim.anosoff2018@yandex.ru", [event.organizer.email])

@login_required
def edit_event(request, event_id):
    """
    View для редактирования мероприятия.
    Видео обрабатывается автоматически при замене файла.
    """
    from django.forms.models import model_to_dict

    # Получаем активную подписку пользователя
    active_subscription = (
        UserPackageSubscription.objects.filter(
            user=request.user,
            is_active=True,
        )
        .select_related("package")
        .first()
    )

    event = get_object_or_404(Event, id=event_id, organizer=request.user)

    if event.has_sold_tickets:
        messages.error(
            request,
            "Редактирование этого мероприятия запрещено, так как на него уже проданы билеты.",
        )
        return redirect("partner:partner_event_list")

    if request.method == "POST":
        # Передаем request.FILES, чтобы обработать загрузку нового видео
        # Определяем текущий пакет для передачи в форму
        current_package = event.package if event.package else (active_subscription.package if active_subscription else None)
        form = EventForm(request.POST, request.FILES, instance=event, current_package=current_package)

        # --- НАЧАЛО БЛОГА ИЗМЕНЕНИЙ ---
        # Удаляем старые файлы, если пришли новые
        new_video_file = request.FILES.get("video_url")
        if new_video_file and event.video_url:
            event.video_url.delete(save=False)

        new_image_file = request.FILES.get("image")
        if new_image_file and event.image:
            event.image.delete(save=False)
        # --- КОНЕЦ БЛОГА ИЗМЕНЕНИЙ ---

        # Очищаем медиафайлы (изображение, программа), если были отмечены флажки в форме
        for field_name in ["image", "program_file"]:
            clear_field_name = f"{field_name}-clear"
            if clear_field_name in request.POST:
                current_file = getattr(event, field_name)
                if current_file:
                    current_file.delete(save=False)
                setattr(event, field_name, None)

        if form.is_valid():
            # Проверяем типы загружаемых файлов
            video_file = request.FILES.get("video_url")
            if video_file:
                # Проверяем расширение видео
                valid_video_extensions = ['.mp4', '.mov', '.avi']
                if not any(video_file.name.lower().endswith(ext) for ext in valid_video_extensions):
                    messages.error(
                        request,
                        "Неверный формат видео. Разрешены только файлы MP4, MOV, AVI"
                    )
                    return render(
                        request,
                        "partner/event_form.html",
                        {
                            "form": form,
                            "is_edit": True,
                            "ticket_data": [
                                {"name": name, "price": price, "quantity": quantity}
                                for name, price, quantity in zip(
                                    request.POST.getlist("ticket_name[]"),
                                    request.POST.getlist("ticket_price[]"),
                                    request.POST.getlist("ticket_quantity[]"),
                                )
                                if name and price and quantity
                            ],
                            "rejection_messages": get_rejection_messages(request),
                            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                            "has_free_tickets": False,
                            "packages": EventPackage.objects.all(),
                        },
                    )

# Проверяем длительность видео
                try:
                    temp_file_path = None
                    temp_file_path = tempfile.mktemp(suffix=os.path.splitext(video_file.name)[1])
                    
                    # Сохраняем видео во временный файл
                    with open(temp_file_path, 'wb+') as temp_file:
                        for chunk in video_file.chunks():
                            temp_file.write(chunk)

                    # Проверяем длительность
                    with VideoFileClip(temp_file_path) as video:
                        duration = video.duration
                        max_duration = 310  # 5 минут в секундах
                        if duration > max_duration:
                            if os.path.exists(temp_file_path):
                                os.unlink(temp_file_path)
                            messages.error(
                                request,
                                "Длительность видео превышает 5 минут. Пожалуйста, загрузите видео не длиннее 5 минут."
                            )
                            return render(
                                request,
                                "partner/event_form.html",
                                {
                                    "form": form,
                                    "is_edit": True,
                                    "ticket_data": [
                                        {"name": name, "price": price, "quantity": quantity}
                                        for name, price, quantity in zip(
                                            request.POST.getlist("ticket_name[]"),
                                            request.POST.getlist("ticket_price[]"),
                                            request.POST.getlist("ticket_quantity[]"),
                                        )
                                        if name and price and quantity
                                    ],
                                    "rejection_messages": get_rejection_messages(request),
                                    "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                                    "has_free_tickets": False,
                                    "packages": EventPackage.objects.all(),
                                },
                            )

                    # Удаляем временный файл
                    if os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
                    
                    # Восстанавливаем указатель файла после проверки
                    video_file.file.seek(0)
                except Exception as e:
                    logger.error(f"Ошибка при проверке длительности видео: {str(e)}")
                    # Очищаем временный файл при ошибке
                    if temp_file_path and os.path.exists(temp_file_path):
                        try:
                            os.unlink(temp_file_path)
                        except:
                            pass
                    messages.error(
                        request,
                        "Произошла ошибка при проверке длительности видео. Пожалуйста, попробуйте еще раз."
                    )
                    return render(
                        request,
                        "partner/event_form.html",
                        {
                            "form": form,
                            "is_edit": True,
                            "ticket_data": [
                                {"name": name, "price": price, "quantity": quantity}
                                for name, price, quantity in zip(
                                    request.POST.getlist("ticket_name[]"),
                                    request.POST.getlist("ticket_price[]"),
                                    request.POST.getlist("ticket_quantity[]"),
                                )
                                if name and price and quantity
                            ],
                            "rejection_messages": get_rejection_messages(request),
                            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                            "has_free_tickets": False,
                            "packages": EventPackage.objects.all(),
                        },
                    )

            pdf_file = request.FILES.get("program_file")
            if pdf_file:
                # Проверяем расширение PDF
                if not pdf_file.name.lower().endswith('.pdf'):
                    messages.error(
                        request,
                        "Неверный формат файла. Разрешены только PDF файлы"
                    )
                    return render(
                        request,
                        "partner/event_form.html",
                        {
                            "form": form,
                            "is_edit": True,
                            "ticket_data": [
                                {"name": name, "price": price, "quantity": quantity}
                                for name, price, quantity in zip(
                                    request.POST.getlist("ticket_name[]"),
                                    request.POST.getlist("ticket_price[]"),
                                    request.POST.getlist("ticket_quantity[]"),
                                )
                                if name and price and quantity
                            ],
                            "rejection_messages": get_rejection_messages(request),
                            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                            "has_free_tickets": False,
                            "packages": EventPackage.objects.all(),
                        },
                    )

            # Проверяем типы загружаемых изображений
            main_image = request.FILES.get("image")
            additional_images = request.FILES.getlist("images")

            # Валидация основного изображения
            if main_image:
                valid_image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
                if not any(main_image.name.lower().endswith(ext) for ext in valid_image_extensions):
                    messages.error(
                        request,
                        "Неверный формат основного изображения. Разрешены только JPG, PNG, GIF, WEBP"
                    )
                    return render(
                        request,
                        "partner/event_form.html",
                        {
                            "form": form,
                            "is_edit": True,
                            "ticket_data": [
                                {"name": name, "price": price, "quantity": quantity}
                                for name, price, quantity in zip(
                                    request.POST.getlist("ticket_name[]"),
                                    request.POST.getlist("ticket_price[]"),
                                    request.POST.getlist("ticket_quantity[]"),
                                )
                                if name and price and quantity
                            ],
                            "rejection_messages": get_rejection_messages(request),
                            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                            "has_free_tickets": False,
                            "packages": EventPackage.objects.all(),
                        },
                    )

            # Валидация дополнительных изображений
            if additional_images:
                valid_image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
                for image in additional_images:
                    if not any(image.name.lower().endswith(ext) for ext in valid_image_extensions):
                        messages.error(
                            request,
                            "Неверный формат дополнительного изображения. Разрешены только JPG, PNG, GIF, WEBP"
                        )
                        return render(
                            request,
                            "partner/event_form.html",
                            {
                                "form": form,
                                "is_edit": True,
                                "ticket_data": [
                                    {"name": name, "price": price, "quantity": quantity}
                                    for name, price, quantity in zip(
                                        request.POST.getlist("ticket_name[]"),
                                        request.POST.getlist("ticket_price[]"),
                                        request.POST.getlist("ticket_quantity[]"),
                                    )
                                    if name and price and quantity
                                ],
                                "rejection_messages": get_rejection_messages(request),
                                "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                                "has_free_tickets": False,
                                "packages": EventPackage.objects.all(),
                            },
                        )

            # Проверяем ограничения пакета на количество фотографий
            main_image = request.FILES.get("image")
            additional_images = request.FILES.getlist("images")

            # Считаем общее количество загружаемых фото
            total_images = 0
            if main_image:
                total_images += 1
            if additional_images:
                total_images += len(additional_images)

            # Считаем уже существующие фото (если не заменяем основное)
            existing_images_count = 0
            if not main_image and event.image:
                existing_images_count += 1
            if not additional_images:
                existing_images_count += event.images.count()

            # Общее количество фото после загрузки
            final_images_count = existing_images_count + total_images

            # Проверяем не превышаем ли лимит пакета
            package = event.package if event.package else (active_subscription.package if active_subscription else None)
            if package and final_images_count > package.max_photos:
                additional_photos_allowed = package.max_photos - 1
                messages.error(
                    request,
                    f"Ваш пакет позволяет загрузить не более {package.max_photos} фотографий "
                    f"(1 основное + {additional_photos_allowed} дополнительных). "
                    f"Вы пытаетесь загрузить {final_images_count} фото."
                )
                return render(
                    request,
                    "partner/event_form.html",
                    {
                        "form": form,
                        "is_edit": True,
                        "ticket_data": [
                            {"name": name, "price": price, "quantity": quantity}
                            for name, price, quantity in zip(
                                request.POST.getlist("ticket_name[]"),
                                request.POST.getlist("ticket_price[]"),
                                request.POST.getlist("ticket_quantity[]"),
                            )
                            if name and price and quantity
                        ],
                        "rejection_messages": get_rejection_messages(request),
                        "main_tags": MainTag.objects.prefetch_related("subtags").all(),
                        "has_free_tickets": False,
                        "packages": EventPackage.objects.all(),
                    },
                )

            # Сохраняем форму
            event = form.save()

            # Основное фото (отдельный input image): если пришло — заменяем
            main_image = request.FILES.get("image")
            if main_image:
                event.image = main_image
                event.save(update_fields=["image"])

            # Удаляем фотографии, которые были отмечены для удаления
            deleted_image_ids = request.POST.get("deleted_image_ids", "")
            if deleted_image_ids:
                from core.models import EventImage
                for image_id in deleted_image_ids.split(","):
                    if image_id:
                        try:
                            image = EventImage.objects.get(id=int(image_id), event=event)
                            image.delete()
                        except EventImage.DoesNotExist:
                            pass

            # Дополнительные фото (input images - list): добавляем новые к существующим
            images = request.FILES.getlist("images")
            if images:
                from core.models import EventImage
                for image in images:
                    EventImage.objects.create(event=event, image=image)

            # Теги
            tags_ids = request.POST.getlist("tags")
            if tags_ids:
                selected_tags = tags_ids[:5]
                event.tags.set(selected_tags)

            # Билеты
            event.tickets.all().delete()
            ticket_names = request.POST.getlist("ticket_name[]")
            ticket_prices = request.POST.getlist("ticket_price[]")
            ticket_quantities = request.POST.getlist("ticket_quantity[]")

            has_free_tickets = False
            has_paid_tickets = False

            for price in ticket_prices:
                if price:
                    try:
                        price_value = (
                            float(price.replace(",", "."))
                            if "," in price
                            else float(price)
                        )
                        if price_value == 0:
                            has_free_tickets = True
                        else:
                            has_paid_tickets = True
                    except (ValueError, TypeError):
                        continue

            if has_free_tickets and has_paid_tickets:
                messages.error(
                    request,
                    "Невозможно создать мероприятие с бесплатными и платными билетами одновременно.",
                )
                return render(
                    request,
                    "partner/event_form.html",
                    {
                        "form": form,
                        "is_edit": True,
                        "ticket_data": [
                            {"name": name, "price": price, "quantity": quantity}
                            for name, price, quantity in zip(
                                ticket_names,
                                ticket_prices,
                                ticket_quantities,
                            )
                            if name and price and quantity
                        ],
                        "rejection_messages": get_rejection_messages(request),
                        "all_tags": Tag.objects.all(),
                    },
                )

            for name, price, quantity in zip(
                ticket_names, ticket_prices, ticket_quantities
            ):
                if name and price and quantity:
                    try:
                        event.tickets.create(
                            name=name,
                            price=(
                                float(price.replace(",", "."))
                                if "," in price
                                else float(price)
                            ),
                            available_quantity=int(quantity),
                        )
                    except (ValueError, TypeError):
                        continue

            return redirect("partner:partner_event_list")

    else:
        # GET: выводим пакет, который сейчас привязан к событию (event.package),
        # а также пакет из активной подписки пользователя (fallback),
        # чтобы информация всегда была доступна как в create_event.
        current_package = event.package if event.package else (active_subscription.package if active_subscription else None)
        form = EventForm(instance=event, current_package=current_package)

        # 1) Пакет события (если заполнен)
        if event.package:
            package = event.package
            active_subscription = (
                UserPackageSubscription.objects.filter(
                    user=request.user,
                    package=package,
                    is_active=True,
                )
                .select_related("package")
                .first()
            )

            package_snapshot = model_to_dict(
                package, fields=[f.name for f in package._meta.fields]
            )
            sub_snapshot = (
                model_to_dict(
                    active_subscription,
                    fields=[f.name for f in active_subscription._meta.fields],
                )
                if active_subscription
                else None
            )

        # 2) Fallback: активная подписка пользователя (как в create_event)
        fallback_subscription = (
            UserPackageSubscription.objects.filter(
                user=request.user,
                is_active=True,
            )
            .select_related("package")
            .first()
        )

        if fallback_subscription:
            fallback_package = fallback_subscription.package
            package_snapshot = model_to_dict(
                fallback_package,
                fields=[f.name for f in fallback_package._meta.fields],
            )
            sub_snapshot = model_to_dict(
                fallback_subscription,
                fields=[f.name for f in fallback_subscription._meta.fields],
            )

    return render(
        request,
        "partner/event_form.html",
        {
            "form": form,
            "is_edit": True,
            "rejection_messages": get_rejection_messages(request),
            "main_tags": MainTag.objects.prefetch_related("subtags").all(),
            "has_free_tickets": False,
            "packages": EventPackage.objects.all(),
        },
    )

@login_required
def partner_event_list(request):
    """
    Отображает список всех мероприятий текущего партнера с возможностью фильтрации.
    """
    # Получаем параметры фильтрации из GET-запроса
    title_query = request.GET.get("title", None)
    date_query = request.GET.get("date", None)

    # Базовый фильтр: только мероприятия текущего пользователя
    events = Event.objects.filter(organizer=request.user)

    # Применяем фильтры
    if title_query:
        events = events.filter(title__icontains=title_query)
    if date_query:
        # Преобразуем строку даты в объект date для фильтрации
        try:
            query_date = datetime.fromisoformat(date_query)
            events = events.filter(date_time__date=query_date)
        except (ValueError, TypeError):
            # Если дата некорректна, игнорируем этот фильтр
            pass

    # Сортируем по дате (новые сверху)
    events = events.order_by("-date_time")

    event_data = []
    for event in events:
        # Суммируем количество проданных билетов по всем типам этого мероприятия
        sold_tickets = sum(
            order.quantity
            for ticket in event.tickets.all()
            for order in ticket.orders.all()
        )
        # Суммируем общее количество доступных билетов
        total = sum(ticket.available_quantity for ticket in event.tickets.all())

        event_data.append(
            {
                "event": event,
                "sold": sold_tickets,
                "total": total,
            }
        )

    context = {
        "events": event_data,
        "rejection_messages": get_rejection_messages(request),
    }
    return render(request, "partner/partner_event_list.html", context)

@login_required
def delete_event(request, event_id):
    """
    Удаляет мероприятие и связанные медиафайлы.
    """
    event = get_object_or_404(Event, id=event_id, organizer=request.user)

    if request.method == "POST":
        # Удаляем медиафайлы, если они существуют
        if event.image:
            event.image.delete()
        if event.video_url:
            event.video_url.delete()
        if event.program_file:
            event.program_file.delete()

        event.delete()
        return redirect("partner:partner_event_list")

    return render(
        request,
        "partner/event_confirm_delete.html",
        {"event": event},
    )

@login_required
def bulk_delete_events(request):
    """
    Удаляет несколько мероприятий за раз.
    """
    if request.method == "POST":
        event_ids = request.POST.getlist("event_ids")
        if not event_ids:
            messages.error(request, "Не выбрано ни одного мероприятия для удаления.")
            return redirect("partner:partner_event_list")

        deleted_count = 0
        for event_id in event_ids:
            try:
                event = Event.objects.get(id=event_id, organizer=request.user)

                # Удаляем медиафайлы через storage backend
                if event.image:
                    event.image.delete()
                if event.video_url:
                    event.video_url.delete()
                if event.program_file:
                    event.program_file.delete()

                # Удаляем объект мероприятия
                event.delete()

                deleted_count += 1
            except Event.DoesNotExist:
                continue
            except Exception as e:
                logger.error(f"Ошибка при удалении мероприятия {event_id}: {str(e)}")
                continue

        messages.success(request, f"Успешно удалено {deleted_count} мероприятий.")
        return redirect("partner:partner_event_list")

    return redirect("partner:partner_event_list")

@login_required
def reports(request):
    """
    Отчеты и статистика продаж для партнера.
    """
    orders = Order.objects.filter(ticket__event__organizer=request.user)

    # Расчет общей статистики (без учёта возвратов)
    non_refunded_orders = orders.exclude(payment_status__in=["canceled", "refunded"])
    total_sales = non_refunded_orders.aggregate(total=Sum("total_price"))["total"] or 0

    # Считаем количество возвратов
    refunded_orders = orders.filter(payment_status__in=["canceled", "refunded"])
    total_refunds = refunded_orders.aggregate(total=Sum("total_price"))["total"] or 0

    # Считаем реальное количество проданных билетов (без возвратов)
    tickets_sold = sum(order.quantity for order in non_refunded_orders)

    # Средний чек = общая выручка / количество проданных билетов
    avg_check = total_sales / tickets_sold if tickets_sold > 0 else 0

    # Получаем данные для графика продаж по дням (без учёта возвратов)
    sales_graph_data = (
        non_refunded_orders.annotate(date=TruncDate("created_at"))
        .values("date")
        .annotate(total=Sum("total_price"))
        .order_by("date")
    )

    # Преобразуем в формат для Chart.js (конвертируем Decimal в float)
    sales_graph_data = {
        item["date"].strftime("%Y-%m-%d"): float(item["total"])
        for item in sales_graph_data
    }

    # Получаем данные об источниках трафика (без учёта возвратов)
    traffic_sources = (
        non_refunded_orders.exclude(utm_source__isnull=True)
        .exclude(utm_source__exact="")
        .values("utm_source")
        .annotate(total=Sum("total_price"), count=Count("id"))
        .order_by("-total")
    )

    # Преобразуем в удобный формат
    traffic_sources_data = {
        item["utm_source"]: {"total": float(item["total"]), "count": item["count"]}
        for item in traffic_sources
    }

    # Получаем список ранее сгенерированных отчётов
    user_reports = SalesReport.objects.filter(partner=request.user).order_by(
        "-created_at"
    )

    # Получаем текущие настройки расписания отчётов
    try:
        report_schedule = ReportSchedule.objects.get(partner=request.user)
    except ReportSchedule.DoesNotExist:
        report_schedule = None

    context = {
        "total_sales": "{:,.2f}".format(total_sales).replace(",", " "),
        "tickets_sold": tickets_sold,
        "avg_check": "{:,.2f}".format(avg_check).replace(",", " "),
        "total_refunds": "{:,.2f}".format(total_refunds).replace(",", " "),
        "refunded_tickets": sum(order.quantity for order in refunded_orders),
        "sales_graph_data": json.dumps(sales_graph_data),
        "traffic_sources_data": traffic_sources_data,
        "user_reports": user_reports,
        "report_schedule": report_schedule,
    }
    return render(request, "partner/reports.html", context)

@login_required
def participant_list(request, event_id):
    """
    Список участников для выбранного мероприятия с поиском, фильтрацией и экспортом.
    """
    # Получаем мероприятие или выдаем 404, если его нет или оно чужое
    event = get_object_or_404(Event, id=event_id, organizer=request.user)

    # Получаем параметры фильтрации из GET-запроса
    search_name = request.GET.get("name", "")
    search_email = request.GET.get("email", "")
    search_status = request.GET.get("status", "")

    # Базовый фильтр: только заказы для этого мероприятия, исключая возвраты
    orders = (
        Order.objects.filter(ticket__event=event)
        .exclude(payment_status__in=["canceled", "refunded"])
        .select_related("ticket")
    )

    # Применяем фильтры
    if search_name:
        orders = orders.filter(participant_data__name__icontains=search_name)
    if search_email:
        orders = orders.filter(participant_data__email__icontains=search_email)
    if search_status == "is_paid":
        orders = orders.filter(is_paid=True)
    elif search_status == "not_paid":
        orders = orders.filter(is_paid=False)

    # Обработка экспорта
    export_format = request.GET.get("export")
    if export_format:
        # Для экспорта также исключаем возвраты
        export_orders = orders.exclude(payment_status__in=["canceled", "refunded"])
        return export_participant_list(export_orders, event, export_format)

    context = {
        "event": event,
        "orders": orders,
    }
    context["rejection_messages"] = get_rejection_messages(request)
    return render(request, "partner/participant_list.html", context)

def export_participant_list(orders, event, export_format):
    """
    Экспортирует список участников в Excel или PDF.
    """
    import io
    from django.http import HttpResponse

    if export_format == "excel":
        # Создаем Excel-файл
        import openpyxl
        from openpyxl.styles import Font, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Участники {event.title}"

        # Заголовки
        headers = [
            "Имя",
            "E-mail",
            "Телефон",
            "Дата покупки",
            "Тип билета",
            "Статус",
            "Цена",
            "Количество билетов",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row_num, order in enumerate(orders, 2):
            ws.cell(
                row=row_num,
                column=1,
                value=f"{order.participant_data.get('first_name', '')} {order.participant_data.get('last_name', '')}".strip(),
            )
            ws.cell(
                row=row_num, column=2, value=order.participant_data.get("email", "")
            )
            ws.cell(
                row=row_num, column=3, value=order.participant_data.get("phone", "")
            )
            ws.cell(
                row=row_num, column=4, value=order.created_at.strftime("%d.%m.%Y %H:%M")
            )
            ws.cell(row=row_num, column=5, value=order.ticket.name)
            ws.cell(
                row=row_num,
                column=6,
                value="Оплачено" if order.is_paid else "Не оплачен",
            )
            ws.cell(row=row_num, column=7, value=f"{order.total_price:.2f} руб.")
            ws.cell(
                row=row_num, column=8, value=f"Количество билетов: {order.quantity}"
            )

        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Отправляем файл
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="Участники_{event.title}.xlsx"'
        )

        wb.save(response)
        return response

    elif export_format == "pdf":
        # Создаем PDF-файл
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Image,
        )
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        import qrcode
        import io

        # Регистрируем шрифт DejaVuSans для поддержки кириллицы
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        font_path = os.path.join(base_dir, "DejaVuSans.ttf")
        pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))

        font_bold_path = os.path.join(base_dir, "DejaVuSans-Bold.ttf")
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", font_bold_path))

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Стили
        styles = getSampleStyleSheet()
        styles["Title"].fontName = "DejaVuSans-Bold"
        styles["Normal"].fontName = "DejaVuSans"

        # Заголовок
        elements.append(Paragraph(f"Список участников: {event.title}", styles["Title"]))

        # Данные для таблицы
        data = [
            [
                "Имя",
                "E-mail",
                "Телефон",
                "Дата покупки",
                "Тип билета",
                "Статус",
                "Цена",
                "Кол-во",
                "QR",
            ]
        ]

        for index, order in enumerate(orders, 1):
            # Генерация QR-кодов
            qr_images = []
            for i in range(order.quantity):
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(f"Order ID: {order.id}, Билет: {i+1}")
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                qr_code_img = io.BytesIO()
                img.save(qr_code_img, format="PNG")
                qr_code_img.seek(0)
                qr_images.append(Image(qr_code_img, width=40, height=40))

            # Для первой строки заказа добавляем QR-коды в таблицу
            if qr_images:
                qr_cell = qr_images[0]  # Первый QR-код в основной строке
                other_qrs = qr_images[
                    1:
                ]  # Остальные QR-коды добавим как дополнительные строки
            else:
                qr_cell = " "
                other_qrs = []

            data.append(
                [
                    f"{order.participant_data.get('first_name', '')} {order.participant_data.get('last_name', '')}".strip(),
                    order.participant_data.get("email", ""),
                    order.participant_data.get("phone", ""),
                    order.created_at.strftime("%d.%m.%Y %H:%M"),
                    order.ticket.name,
                    "Оплачено" if order.is_paid else "Не оплачен",
                    f"{order.total_price:.2f} руб.",
                    f"{order.quantity}",
                    qr_cell,
                ]
            )

            # Добавляем дополнительные QR-коды как отдельные строки в таблицу
            for qr_img in other_qrs:
                data.append(["", "", "", "", "", "", "", "", qr_img])

        # Создаем таблицу
        table = Table(data)
        # Устанавливаем ширину столбцов
        column_widths = [80, 100, 60, 70, 70, 60, 60, 40, 50]
        table._argW = column_widths

        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
                    ("FONTSIZE", (0, 1), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("WORDWRAP", (0, 0), (-1, -1), True),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("ALIGN", (7, 0), (7, -1), "CENTER"),
                ]
            )
        )

        elements.append(table)
        doc.build(elements)

        # Отправляем файл
        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="Участники_{event.title}_с_QR.pdf"'
        )
        return response

    return HttpResponse("Неверный формат экспорта", status=400)

@login_required
def mark_attendance(request, event_id, order_id):
    """
    Отмечает участника как посетившего мероприятие.
    """
    # Получаем мероприятие или выдаем 404, если его нет или оно чужое
    event = get_object_or_404(Event, id=event_id, organizer=request.user)

    # Получаем заказ
    order = get_object_or_404(Order, id=order_id, ticket__event=event)

    if request.method == "POST":
        # Инвертируем статус посещения
        order.attended = not order.attended
        order.save()
        messages.success(
            request,
            f"Статус посещения для {order.participant_data.get('name', 'участника')} обновлен!",
        )

    return redirect("partner:participant_list", event_id=event.id)

@login_required
def check_ticket(request, order_id):
    """
    Проверка билета по QR-коду.
    """
    order = get_object_or_404(Order, id=order_id)

    context = {
        "order": order,
        "is_valid": True,
    }
    return render(request, "partner/ticket_check.html", context)

@login_required
def finances(request):
    orders = Order.objects.filter(ticket__event__organizer=request.user)

    # Считаем общую выручку (только оплаченные заказы, кроме возвратов)
    total_revenue = (
        orders.filter(is_paid=True)
        .exclude(payment_status="refunded")
        .aggregate(total=Sum("total_price"))["total"]
        or 0
    )

    commission_sum = (
        orders.filter(is_paid=True)
        .exclude(payment_status="refunded")
        .annotate(
            event_commission=ExpressionWrapper(
                F("total_price") * (F("ticket__event__commission_rate") / 100),
                output_field=DecimalField(),
            )
        )
        .aggregate(total_commission=Sum("event_commission"))["total_commission"]
        or 0
    )
    commission_amount = commission_sum
    # Сумма к выплате: выручка минус комиссия
    payout_amount = total_revenue - commission_sum

    payout_history = PayoutRequest.objects.filter(organizer=request.user).order_by(
        "-created_at"
    )

    # Получаем реквизиты партнёра
    partner_payout_details = PayoutDetails.objects.filter(partner=request.user)

    context = {
        "total_revenue": total_revenue,
        "commission_amount": commission_amount,
        "payout_amount": payout_amount,
        "payout_history": payout_history,
        "partner_payout_details": partner_payout_details,
    }
    context["rejection_messages"] = get_rejection_messages(request)
    return render(request, "partner/finances.html", context)

@require_POST
@csrf_exempt
def request_payout(request):
    """
    Обработка AJAX-запроса на создание запроса выплаты.
    """
    try:
        data = request.POST
        amount = float(data.get("amount", 0))
        payout_details_id = data.get("payout_details")
        comment = data.get("comment", "")

        if not payout_details_id:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Пожалуйста, выберите реквизиты для выплаты",
                },
                status=400,
            )

        try:
            payout_details = PayoutDetails.objects.get(
                id=payout_details_id, partner=request.user
            )
        except ObjectDoesNotExist:
            return JsonResponse(
                {"status": "error", "message": "Выбранные реквизиты не найдены"},
                status=404,
            )

        # Получаем доступную для выплаты сумму (только оплаченные заказы, кроме возвратов)
        orders = Order.objects.filter(ticket__event__organizer=request.user)
        total_revenue = (
            orders.filter(is_paid=True)
            .exclude(payment_status="refunded")
            .aggregate(total=Sum("total_price"))["total"]
            or 0
        )
        commission_sum = (
            orders.filter(is_paid=True)
            .exclude(payment_status="refunded")
            .annotate(
                event_commission=ExpressionWrapper(
                    F("total_price") * (F("ticket__event__commission_rate") / 100),
                    output_field=DecimalField(),
                )
            )
            .aggregate(total_commission=Sum("event_commission"))["total_commission"]
            or 0
        )
        payout_amount = total_revenue - commission_sum

        # Серверная валидация суммы выплаты
        if amount > payout_amount:
            return JsonResponse(
                {
                    "status": "error",
                    "message": f"Сумма выплаты не может превышать доступную сумму: {payout_amount:.2f} ₽",
                },
                status=400,
            )

        # Создаём запрос на выплату
        payout_request = PayoutRequest.objects.create(
            organizer=request.user,
            amount=amount,
            payment_details=payout_details,
            comment=comment,
            status="pending",
        )

        return JsonResponse(
            {"status": "success", "message": "Запрос на выплату успешно создан!"}
        )

    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": f"Произошла ошибка: {str(e)}"}, status=500
        )

@require_POST
@login_required
def cancel_payout(request, payout_id):
    """
    Отмена заявки на выплату через AJAX.
    """
    try:
        # Ищем заявку, которая принадлежит текущему пользователю
        payout_request = get_object_or_404(
            PayoutRequest, id=payout_id, organizer=request.user
        )

        # Проверяем, можно ли отменить (статус должен быть 'pending')
        if payout_request.status != "pending":
            return JsonResponse(
                {
                    "status": "error",
                    "message": 'Отменить можно только заявку в статусе "Ожидает обработки"',
                },
                status=400,
            )

        # Меняем статус и сохраняем
        payout_request.status = "cancelled"
        payout_request.save()

        return JsonResponse(
            {"status": "success", "message": "Заявка на выплату отменена!"}
        )

    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": f"Произошла ошибка: {str(e)}"}, status=500
        )

@require_POST
@csrf_exempt
def delete_reports(request):
    """
    Удаление выбранных отчётов через AJAX.
    """
    try:
        data = json.loads(request.body)
        report_ids = data.get("report_ids", [])

        if not report_ids:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Не выбрано ни одного отчёта для удаления",
                },
                status=400,
            )

        # Удаляем отчёты и их файлы
        deleted_count = 0
        for report_id in report_ids:
            try:
                report = SalesReport.objects.get(id=report_id, partner=request.user)
                if report.file_path:
                    report.file_path.delete()
                report.delete()
                deleted_count += 1
            except ObjectDoesNotExist:
                continue

        return JsonResponse(
            {
                "status": "success",
                "message": f"Успешно удалено {deleted_count} отчёт(ов)",
            }
        )

    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": f"Произошла ошибка: {str(e)}"}, status=500
        )

@login_required
def payout_details(request):
    """
    Страница для добавления и просмотра реквизитов для выплат.
    """
    # Получаем все реквизиты текущего пользователя
    details = PayoutDetails.objects.filter(partner=request.user)

    if request.method == "POST":
        form = PayoutDetailsForm(request.POST)
        if form.is_valid():
            payout_detail = form.save(commit=False)
            payout_detail.partner = request.user
            payout_detail.save()
            messages.success(request, "Реквизиты успешно сохранены!")
            return redirect("partner:payout_details")
    else:
        form = PayoutDetailsForm()

    context = {
        "form": form,
        "details": details,
        "rejection_messages": get_rejection_messages(request),
    }
    return render(request, "partner/payout_details.html", context)

@login_required
def profile_edit(request):
    """
    View для редактирования профиля партнера.
    Включает обработку видео-визитки.
    """
    from core.models import PartnerDocument
    
    if request.method == "POST":
        # Инициализируем форму с данными и файлами
        user_form = PartnerProfileForm(
            request.POST, request.FILES, instance=request.user
        )

        # --- НАЧАЛО БЛОКА ИЗМЕНЕНИЙ ---
        # Эта логика удаляет старое видео-визитку с диска, если был загружен новый файл.
        # Она должна выполняться ДО валидации и сохранения формы.

        # Проверяем, был ли загружен НОВЫЙ файл для поля 'video_business_card'
        new_video_file = request.FILES.get("video_business_card")

        # Если новый файл есть, и у пользователя уже было старое видео...
        if new_video_file and request.user.video_business_card:
            # ...то удаляем старый файл с диска.
            # Метод .delete() у FileField удаляет файл с диска.
            # Параметр save=False важен: мы не хотим сохранять модель сейчас.
            request.user.video_business_card.delete(save=False)
        # --- КОНЕЦ БЛОКА ИЗМЕНЕНИЙ ---

        # Обработка основной формы профиля (включая видео-визитку)
        if user_form.is_valid():
            user_form.save()  # Сохранение здесь запустит сигнал для обработки нового видео

        # Обработка формы смены пароля
        password_form = PasswordChangeForm(user=request.user, data=request.POST)
        if password_form.is_valid():
            password_form.save()
            update_session_auth_hash(request, password_form.user)

        # Обработка формы загрузки документов
        if "upload_documents" in request.POST:
            document_form = DocumentUploadForm(
                request.POST, request.FILES, user=request.user
            )
            if document_form.is_valid():
                # Если был статус rejected, удаляем старый документ
                if request.user.verification_status == "rejected":
                    old_doc = PartnerDocument.objects.filter(
                        user=request.user, 
                        is_approved=False
                    ).first()
                    if old_doc and old_doc.document:
                        # Удаляем физический файл
                        try:
                            old_doc.document.delete(save=False)
                        except Exception as e:
                            logger.error(f"Ошибка при удалении старого документа: {e}")
                        # Удаляем запись из БД
                        old_doc.delete()

                document = document_form.save()
                request.user.verification_status = "pending"
                request.user.save()
                messages.success(request, "Ваши документы загружены и находятся на рассмотрении.")
            else:
                messages.error(request, "Ошибка при загрузке документов. Пожалуйста, исправьте ошибки ниже.")

        messages.success(request, "Ваши изменения успешно сохранены!")
        return redirect("partner:dashboard")

    else:
        user_form = PartnerProfileForm(instance=request.user)
        password_form = PasswordChangeForm(user=request.user)
        document_form = DocumentUploadForm(user=request.user)

    # Если форма документов была отправлена с ошибками, передаём её в контекст
    if request.method == "POST" and "upload_documents" in request.POST:
        document_form = DocumentUploadForm(request.POST, request.FILES, user=request.user)

    # Получаем последний отклонённый документ для отображения причины
    last_rejected_doc = PartnerDocument.objects.filter(
        user=request.user, 
        is_approved=False,
        rejection_reason__isnull=False
    ).first()

    context = {
        "user_form": user_form,
        "password_form": password_form,
        "document_form": document_form,
        "rejection_messages": get_rejection_messages(request),
        "last_rejected_document": last_rejected_doc,
    }
    return render(request, "partner/profile_edit.html", context)

@login_required
def generate_report(request):
    """
    Генерирует отчёт о продажах за указанный период в выбранном формате.
    """
    if request.method == "POST":
        period_start = request.POST.get("period_start")
        period_end = request.POST.get("period_end")
        report_type = request.POST.get("report_type")
        send_email = request.POST.get("send_email", "false").lower() == "true"

        try:
            period_start = datetime.strptime(period_start, "%Y-%m-%d").date()
            period_end = datetime.strptime(period_end, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse(
                {"status": "error", "message": "Неверный формат даты"},
                status=400,
            )

        try:
            # Генерируем отчёт
            report_file = generate_sales_report(
                request.user, period_start, period_end, report_type
            )

            # Сохраняем отчёт в модели
            report = SalesReport.objects.create(
                partner=request.user,
                period_start=period_start,
                period_end=period_end,
                report_type=report_type,
                status="completed",
            )

            # Сохраняем файл
            if report_type == "csv":
                file_name = f"report_{period_start}_{period_end}.csv"
                report.file_path.save(
                    file_name,
                    ContentFile(report_file.read()),
                )
            elif report_type == "excel":
                file_name = f"report_{period_start}_{period_end}.xlsx"
                report.file_path.save(
                    file_name,
                    ContentFile(report_file.getvalue()),
                )
            else:
                file_name = f"report_{period_start}_{period_end}.pdf"
                report.file_path.save(
                    file_name,
                    ContentFile(report_file.getvalue()),
                )

            # Если нужно отправить на email
            if send_email:
                email = EmailMessage(
                    subject=f"Отчёт о продажах с {period_start} по {period_end}",
                    body=f"Добрый день!\n\nПрикрепляем отчёт о продажах за период с {period_start} по {period_end}.\n\nС уважением, ваша платформа мероприятий.",
                    from_email=None,
                    to=[request.user.email],
                )
                email.attach(
                    file_name, report_file.getvalue(), f"application/{report_type}"
                )
                email.send()

            # Возвращаем ссылку на скачивание
            return JsonResponse(
                {
                    "status": "success",
                    "download_url": report.file_path.url,
                    "email_sent": send_email,
                }
            )

        except Exception as e:
            return JsonResponse(
                {"status": "error", "message": str(e)},
                status=500,
            )

    return JsonResponse(
        {"status": "error", "message": "Неверный метод запроса"},
        status=405,
    )

@login_required
def report_schedule(request):
    """
    Настройка расписания отправки отчётов.
    """
    try:
        # Получаем или создаём настройки расписания для текущего пользователя
        schedule, created = ReportSchedule.objects.get_or_create(partner=request.user)

        if request.method == "POST":
            form = ReportScheduleForm(
                request.POST, instance=schedule, partner=request.user
            )
            if form.is_valid():
                form.save()
                messages.success(request, "Настройки расписания успешно сохранены!")
                return redirect("partner:report_schedule")
            else:
                messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
        else:
            form = ReportScheduleForm(instance=schedule, partner=request.user)

        return render(
            request,
            "partner/report_schedule.html",
            {"form": form, "rejection_messages": get_rejection_messages(request)},
        )
    except Exception as e:
        import logging

        logger = logging.getLogger(__name__)
        logger.exception("Ошибка в представлении report_schedule: %s", str(e))
        messages.error(request, f"Произошла ошибка: {str(e)}")
        return redirect("partner:dashboard")

@login_required
def remove_media(request, media_type, media_id):
    """
    View для удаления медиафайлов через AJAX.
    media_id - это ID мероприятия (event_id) для image, video_url, program_file
    """
    import logging
    logger = logging.getLogger(__name__)
    
    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "Method not allowed"}, status=405
        )

    try:
        logger.info(f"remove_media: media_type={media_type}, media_id={media_id}, user={request.user}")
        
        if media_type in ["image", "video_url", "program_file"]:
            event = Event.objects.get(id=media_id, organizer=request.user)
            logger.info(f"remove_media: найдено мероприятие {event.id}")
            
            if media_type == "image" and event.image:
                logger.info(f"remove_media: удаляем image={event.image}")
                event.delete_file_field("image")
                event.image = None
                event.save()
                logger.info(f"remove_media: image успешно удалён")
                return JsonResponse({"status": "success"})
            elif media_type == "video_url" and event.video_url:
                logger.info(f"remove_media: удаляем video_url={event.video_url}")
                event.delete_file_field("video_url")
                event.video_url = None
                event.save()
                logger.info(f"remove_media: video_url успешно удалён")
                return JsonResponse({"status": "success"})
            elif media_type == "program_file" and event.program_file:
                logger.info(f"remove_media: удаляем program_file={event.program_file}")
                event.delete_file_field("program_file")
                event.program_file = None
                event.save()
                logger.info(f"remove_media: program_file успешно удалён")
                return JsonResponse({"status": "success"})
            else:
                logger.warning(f"remove_media: файл {media_type} не найден у мероприятия {media_id}")
                return JsonResponse(
                    {"status": "error", "message": "Media not found"}, status=404
                )

        elif media_type == "video_business_card":
            if request.user.video_business_card:
                request.user.delete_file_field("video_business_card")
                request.user.video_business_card = None
                request.user.save()
                return JsonResponse({"status": "success"})

        return JsonResponse(
            {"status": "error", "message": "Media not found"}, status=404
        )
    except Event.DoesNotExist:
        logger.error(f"remove_media: мероприятие {media_id} не найдено")
        return JsonResponse(
            {"status": "error", "message": "Event not found"}, status=404
        )
    except Exception as e:
        logger.error(f"remove_media: ошибка: {e}", exc_info=True)
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

@login_required
def change_password(request):
    """Отдельная страница для смены пароля в личном кабинете партнёра."""
    if request.method == "POST":
        password_form = PasswordChangeForm(user=request.user, data=request.POST)
        if password_form.is_valid():
            password_form.save()
            update_session_auth_hash(request, password_form.user)
            messages.success(request, "Пароль успешно изменён!")
            return redirect("partner:dashboard")
    else:
        password_form = PasswordChangeForm(user=request.user)

    return render(
        request,
        "change_password.html",
        {"form": password_form, "rejection_messages": get_rejection_messages(request)},
    )

@login_required
def remove_event_image(request, image_id):
    """Удаление фотографии мероприятия через AJAX."""
    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "Method not allowed"}, status=405
        )

    try:
        from core.models import EventImage

        image = EventImage.objects.get(id=image_id, event__organizer=request.user)
        image.delete_file_field("image")  # Корректное удаление из S3
        image.delete()  # Удаляем запись из БД
        return JsonResponse({"status": "success"})
    except EventImage.DoesNotExist:
        return JsonResponse(
            {"status": "error", "message": "Image not found"}, status=404
        )
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

def send_partner_all_tickets_sold_notification(event):
    """
    Отправляет уведомление партнёру о том, что все билеты на мероприятие выкуплены.
    """
    subject = f"Все билеты на мероприятие '{event.title}' выкуплены"
    organizer_email = event.organizer.email

    # Формируем сообщение
    message = f"""
    <html>
    <body>
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <h2 style="color: #2c3e50;">Здравствуйте, {event.organizer.get_full_name()}!</h2>

            <p>Поздравляем! Все билеты на ваше мероприятие <strong>{event.title}</strong> были успешно выкуплены.</p>

            <div style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <p><strong>Дата и время:</strong> {event.date_time.strftime('%d.%m.%Y %H:%M')}</p>
                <p><strong>Место проведения:</strong> {event.get_place_address}</p>
            </div>

            <p>Теперь вы можете подготовиться к проведению мероприятия.</p>

            <div style="
                margin-top: 40px;
                padding-top: 20px;
                border-top: 1px solid #eee;
                font-size: 12px;
                color: #7f8c8d;
            ">
                <p>Если у вас возникли вопросы, обратитесь в нашу <a href="#">службу поддержки</a>.</p>
            </div>
        </div>
    </body>
    </html>
    """
    send_mail(
        subject,
        "",
        "dim.anosoff2018@yandex.ru",
        [organizer_email],
        html_message=message,
    )