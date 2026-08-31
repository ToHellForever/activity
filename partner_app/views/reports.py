"""Отчёты, список участников и экспорт."""
import json
import logging
from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.files.base import ContentFile
from django.core.mail import EmailMessage
from django.core.exceptions import ObjectDoesNotExist
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate

from core.models import Event, Order, OrderTicket
from ..forms import ReportScheduleForm
from ..models import SalesReport, ReportSchedule
from ..utils import generate_sales_report
from .decorators import check_partner_status, get_rejection_messages

logger = logging.getLogger(__name__)


@login_required
@check_partner_status('can_request_reports')
def reports(request):
    orders = Order.objects.filter(ticket__event__organizer=request.user)

    # Расчет общей статистики (без учёта возвратов)
    non_refunded_orders = orders.exclude(payment_status__in=["canceled", "refunded"])
    total_sales = non_refunded_orders.aggregate(total=Sum("total_price"))["total"] or 0

    # Считаем количество возвратов по заказам
    refunded_orders = orders.filter(payment_status__in=["canceled", "refunded"])
    total_refunds = refunded_orders.aggregate(total=Sum("total_price"))["total"] or 0

    # Считаем реальное количество проданных билетов (без возвратов на уровне заказа)
    tickets_sold = sum(order.quantity for order in non_refunded_orders)

    # --- Считаем отдельно возвращённые билеты через OrderTicket ---
    # Билеты возвращённые поштучно (is_refunded=True) в заказах, которые сами не refunded
    refunded_order_tickets = OrderTicket.objects.filter(
        order__ticket__event__organizer=request.user,
        is_refunded=True,
    ).exclude(
        order__payment_status__in=["canceled", "refunded"]  # чтобы не дублировать
    )

    refunded_order_tickets_count = refunded_order_tickets.count()

    # Сумма возвратов по отдельным билетам
    refunded_order_tickets_sum = (
        refunded_order_tickets
        .aggregate(total=Sum("order__ticket__price"))["total"] or 0
    )

    # Итоговые цифры возвратов (заказы + отдельные билеты)
    total_refunds = float(total_refunds) + float(refunded_order_tickets_sum)
    total_refunded_tickets = (
        sum(order.quantity for order in refunded_orders) + refunded_order_tickets_count
    )

    # Вычитаем поштучно возвращённые из проданных
    tickets_sold = tickets_sold - refunded_order_tickets_count

    # Вычитаем поштучно возвращённые из суммы продаж
    total_sales = float(total_sales) - float(refunded_order_tickets_sum)

    # Средний чек
    avg_check = total_sales / tickets_sold if tickets_sold > 0 else 0

    # График продаж по дням (без учёта возвратов)
    sales_graph_data = (
        non_refunded_orders.annotate(date=TruncDate("created_at"))
        .values("date")
        .annotate(total=Sum("total_price"))
        .order_by("date")
    )
    sales_graph_data = {
        item["date"].strftime("%Y-%m-%d"): float(item["total"])
        for item in sales_graph_data
    }

    # Источники трафика
    traffic_sources = (
        non_refunded_orders.exclude(utm_source__isnull=True)
        .exclude(utm_source__exact="")
        .values("utm_source")
        .annotate(total=Sum("total_price"), count=Count("id"))
        .order_by("-total")
    )
    traffic_sources_data = {
        item["utm_source"]: {"total": float(item["total"]), "count": item["count"]}
        for item in traffic_sources
    }

    # Отчёты и расписание
    user_reports = SalesReport.objects.filter(partner=request.user).order_by("-created_at")
    try:
        report_schedule = ReportSchedule.objects.get(partner=request.user)
    except ReportSchedule.DoesNotExist:
        report_schedule = None

    partner_profile = getattr(request.user, 'partner_profile', None)

    context = {
        "total_sales": "{:,.2f}".format(total_sales).replace(",", " "),
        "tickets_sold": tickets_sold,
        "avg_check": "{:,.2f}".format(avg_check).replace(",", " "),
        "total_refunds": "{:,.2f}".format(total_refunds).replace(",", " "),
        "refunded_tickets": total_refunded_tickets,  # учитывает оба типа возвратов
        "sales_graph_data": json.dumps(sales_graph_data),
        "traffic_sources_data": traffic_sources_data,
        "user_reports": user_reports,
        "report_schedule": report_schedule,
        "partner_profile": partner_profile,
    }
    return render(request, "partner/reports.html", context)


@login_required
def participant_list(request, event_id):
    """
    Список участников для выбранного мероприятия с поиском, фильтрацией и экспортом.
    """
    # Получаем мероприятие или выдаем 404, если его нет или оно чужое
    event = get_object_or_404(Event, id=event_id, organizer=request.user)

    # Получаем параметры фильтрации из GET-запроса
    search_name = request.GET.get("name", "")
    search_email = request.GET.get("email", "")
    search_status = request.GET.get("status", "")

    # Базовый фильтр: только заказы для этого мероприятия, исключая возвраты
    orders = (
        Order.objects.filter(ticket__event=event)
        .exclude(payment_status__in=["canceled", "refunded"])
        .select_related("ticket")
        .prefetch_related("tickets")
    )

    # Применяем фильтры
    if search_name:
        orders = orders.filter(participant_data__name__icontains=search_name)
    if search_email:
        orders = orders.filter(participant_data__email__icontains=search_email)
    if search_status == "is_paid":
        orders = orders.filter(is_paid=True)
    elif search_status == "not_paid":
        orders = orders.filter(is_paid=False)

    # Обработка экспорта
    export_format = request.GET.get("export")
    if export_format:
        # Для экспорта также исключаем возвраты
        export_orders = orders.exclude(payment_status__in=["canceled", "refunded"])
        return export_participant_list(export_orders, event, export_format)

    partner_profile = getattr(request.user, 'partner_profile', None)
    context = {
        "event": event,
        "orders": orders,
        "partner_profile": partner_profile,
    }
    context["rejection_messages"] = get_rejection_messages(request)
    return render(request, "partner/participant_list.html", context)


def export_participant_list(orders, event, export_format):
    """
    Экспортирует список участников в Excel или PDF.
    """
    if export_format == "excel":
        # Создаем Excel-файл
        import openpyxl
        from openpyxl.styles import Font, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Участники {event.title}"

        # Заголовки
        headers = [
            "Имя",
            "E-mail",
            "Телефон",
            "Дата покупки",
            "Тип билета",
            "Статус",
            "Цена",
            "Билет",
            "Посетил",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row_num, order in enumerate(orders, 2):
            ws.cell(
                row=row_num,
                column=1,
                value=f"{order.participant_data.get('first_name', '')} {order.participant_data.get('last_name', '')}".strip(),
            )
            ws.cell(
                row=row_num, column=2, value=order.participant_data.get("email", "")
            )
            ws.cell(
                row=row_num, column=3, value=order.participant_data.get("phone", "")
            )
            ws.cell(
                row=row_num, column=4, value=order.created_at.strftime("%d.%m.%Y %H:%M")
            )
            ws.cell(row=row_num, column=5, value=order.ticket.name)
            ws.cell(
                row=row_num,
                column=6,
                value="Оплачено" if order.is_paid else "Не оплачен",
            )
            ws.cell(row=row_num, column=7, value=f"{order.total_price:.2f} руб.")
            ws.cell(
                row=row_num, column=8, value=f"Билетов: {order.quantity}"
            )
            # Статус посещения — проверяем по OrderTicket
            attended_tickets = order.tickets.filter(attended=True).count()
            total_tickets = order.tickets.count()
            if attended_tickets == 0:
                visited = "Нет"
            elif attended_tickets == total_tickets:
                visited = "Да"
            else:
                visited = f"Частично ({attended_tickets}/{total_tickets})"
            ws.cell(row=row_num, column=9, value=visited)

        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Отправляем файл
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="Участники_{event.title}.xlsx"'
        )

        wb.save(response)
        return response

    elif export_format == "pdf":
        # Создаем PDF-файл
        import io
        import os

        import qrcode
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Image,
        )
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        # Регистрируем шрифт DejaVuSans для поддержки кириллицы
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        font_path = os.path.join(base_dir, "DejaVuSans.ttf")
        pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))

        font_bold_path = os.path.join(base_dir, "DejaVuSans-Bold.ttf")
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", font_bold_path))

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Стили
        styles = getSampleStyleSheet()
        styles["Title"].fontName = "DejaVuSans-Bold"
        styles["Normal"].fontName = "DejaVuSans"

        # Заголовок
        elements.append(Paragraph(f"Список участников: {event.title}", styles["Title"]))

        # Данные для таблицы
        data = [
            [
                "Имя",
                "E-mail",
                "Телефон",
                "Дата покупки",
                "Тип билета",
                "Статус",
                "Цена",
                "Кол-во",
                "QR",
            ]
        ]

        for index, order in enumerate(orders, 1):
            # Генерация QR-кодов
            qr_images = []
            for i in range(order.quantity):
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(f"Order ID: {order.id}, Билет: {i+1}")
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                qr_code_img = io.BytesIO()
                img.save(qr_code_img, format="PNG")
                qr_code_img.seek(0)
                qr_images.append(Image(qr_code_img, width=40, height=40))

            # Для первой строки заказа добавляем QR-коды в таблицу
            if qr_images:
                qr_cell = qr_images[0]  # Первый QR-код в основной строке
                other_qrs = qr_images[
                    1:
                ]  # Остальные QR-коды добавим как дополнительные строки
            else:
                qr_cell = " "
                other_qrs = []

            data.append(
                [
                    f"{order.participant_data.get('first_name', '')} {order.participant_data.get('last_name', '')}".strip(),
                    order.participant_data.get("email", ""),
                    order.participant_data.get("phone", ""),
                    order.created_at.strftime("%d.%m.%Y %H:%M"),
                    order.ticket.name,
                    "Оплачено" if order.is_paid else "Не оплачен",
                    f"{order.total_price:.2f} руб.",
                    f"{order.quantity}",
                    qr_cell,
                ]
            )

            # Добавляем дополнительные QR-коды как отдельные строки в таблицу
            for qr_img in other_qrs:
                data.append(["", "", "", "", "", "", "", "", qr_img])

        # Создаем таблицу
        table = Table(data)
        # Устанавливаем ширину столбцов
        column_widths = [80, 100, 60, 70, 70, 60, 60, 40, 50]
        table._argW = column_widths

        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
                    ("FONTSIZE", (0, 1), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("WORDWRAP", (0, 0), (-1, -1), True),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("ALIGN", (7, 0), (7, -1), "CENTER"),
                ]
            )
        )

        elements.append(table)
        doc.build(elements)

        # Отправляем файл
        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="Участники_{event.title}_с_QR.pdf"'
        )
        return response

    return HttpResponse("Неверный формат экспорта", status=400)


@login_required
def mark_attendance(request, event_id, order_id, ticket_number=1):
    """
    Отмечает конкретный билет в заказе как посещённый.
    """
    # Получаем мероприятие или выдаем 404, если его нет или оно чужое
    event = get_object_or_404(Event, id=event_id, organizer=request.user)

    # Получаем заказ
    order = get_object_or_404(Order, id=order_id, ticket__event=event)

    # Получаем конкретный билет в заказе
    order_ticket = get_object_or_404(
        OrderTicket, order=order, ticket_number=ticket_number
    )

    if request.method == "POST":
        # Инвертируем статус посещения конкретного билета
        order_ticket.attended = not order_ticket.attended
        order_ticket.save()

        # Если AJAX — возвращаем JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'attended': order_ticket.attended,
                'ticket_number': order_ticket.ticket_number,
                'order_id': order.id,
                'event_id': event.id,
                'payment_status': order.payment_status,
            })

        messages.success(
            request,
            f"Билет #{order_ticket.ticket_number} в заказе #{order.id} обновлён!",
        )
        return redirect("partner:participant_list", event_id=event.id)

    # Для GET-запроса возвращаемся к списку участников
    return redirect("partner:participant_list", event_id=event.id)


@login_required
@check_partner_status('can_request_reports')
def generate_report(request):
    """
    Генерирует отчёт о продажах за указанный период в выбранном формате.
    """
    if request.method == "POST":
        period_start = request.POST.get("period_start")
        period_end = request.POST.get("period_end")
        report_type = request.POST.get("report_type")
        send_email = request.POST.get("send_email", "false").lower() == "true"

        try:
            period_start = datetime.strptime(period_start, "%Y-%m-%d").date()
            period_end = datetime.strptime(period_end, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse(
                {"status": "error", "message": "Неверный формат даты"},
                status=400,
            )

        try:
            # Генерируем отчёт
            report_file = generate_sales_report(
                request.user, period_start, period_end, report_type
            )

            # Сохраняем отчёт в модели
            report = SalesReport.objects.create(
                partner=request.user,
                period_start=period_start,
                period_end=period_end,
                report_type=report_type,
                status="completed",
            )

            # Сохраняем файл
            if report_type == "csv":
                file_name = f"report_{period_start}_{period_end}.csv"
                report.file_path.save(
                    file_name,
                    ContentFile(report_file.read()),
                )
            elif report_type == "excel":
                file_name = f"report_{period_start}_{period_end}.xlsx"
                report.file_path.save(
                    file_name,
                    ContentFile(report_file.getvalue()),
                )
            else:
                file_name = f"report_{period_start}_{period_end}.pdf"
                report.file_path.save(
                    file_name,
                    ContentFile(report_file.getvalue()),
                )

            # Если нужно отправить на email
            if send_email:
                email = EmailMessage(
                    subject=f"Отчёт о продажах с {period_start} по {period_end}",
                    body=f"Добрый день!\n\nПрикрепляем отчёт о продажах за период с {period_start} по {period_end}.\n\nС уважением, ваша платформа мероприятий.",
                    from_email=None,
                    to=[request.user.email],
                )
                email.attach(
                    file_name, report_file.getvalue(), f"application/{report_type}"
                )
                email.send()

            # Возвращаем ссылку на скачивание
            return JsonResponse(
                {
                    "status": "success",
                    "download_url": report.file_path.url,
                    "email_sent": send_email,
                }
            )

        except Exception as e:
            return JsonResponse(
                {"status": "error", "message": str(e)},
                status=500,
            )

    return JsonResponse(
        {"status": "error", "message": "Неверный метод запроса"},
        status=405,
    )


@login_required
@check_partner_status('can_request_reports')
def report_schedule(request):
    """
    Настройка расписания отправки отчётов.
    """
    try:
        # Получаем или создаём настройки расписания для текущего пользователя
        schedule, created = ReportSchedule.objects.get_or_create(partner=request.user)

        if request.method == "POST":
            form = ReportScheduleForm(
                request.POST, instance=schedule, partner=request.user
            )
            if form.is_valid():
                form.save()
                messages.success(request, "Настройки расписания успешно сохранены!")
                return redirect("partner:report_schedule")
            else:
                messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
        else:
            form = ReportScheduleForm(instance=schedule, partner=request.user)

        partner_profile = getattr(request.user, 'partner_profile', None)
        return render(
            request,
            "partner/report_schedule.html",
            {"form": form, "rejection_messages": get_rejection_messages(request), "partner_profile": partner_profile},
        )
    except Exception as e:
        logger.exception("Ошибка в представлении report_schedule: %s", str(e))
        messages.error(request, f"Произошла ошибка: {str(e)}")
        return redirect("partner:dashboard")


@require_POST
@csrf_exempt
def delete_reports(request):
    """
    Удаление выбранных отчётов через AJAX.
    """
    try:
        data = json.loads(request.body)
        report_ids = data.get("report_ids", [])

        if not report_ids:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Не выбрано ни одного отчёта для удаления",
                },
                status=400,
            )

        # Удаляем отчёты и их файлы
        deleted_count = 0
        for report_id in report_ids:
            try:
                report = SalesReport.objects.get(id=report_id, partner=request.user)
                if report.file_path:
                    report.file_path.delete()
                report.delete()
                deleted_count += 1
            except ObjectDoesNotExist:
                continue

        return JsonResponse(
            {
                "status": "success",
                "message": f"Успешно удалено {deleted_count} отчёт(ов)",
            }
        )

    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": f"Произошла ошибка: {str(e)}"}, status=500
        )
